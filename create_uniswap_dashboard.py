#!/usr/bin/env python3
"""
Скрипт для создания Dashboard системы учёта пулов ликвидности Uniswap
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def create_uniswap_dashboard():
    """Создаёт Excel файл с Dashboard для учёта пулов Uniswap"""

    wb = Workbook()

    # Удаляем стандартный лист
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Создаём листы
    create_pools_sheet(wb)
    create_profit_analysis_sheet(wb)
    create_dashboard_sheet(wb)

    # Сохраняем файл
    filename = 'Uniswap_Pool_Dashboard.xlsx'
    wb.save(filename)
    print(f"✅ Файл {filename} успешно создан!")
    return filename


def create_pools_sheet(wb):
    """Создаёт лист 'Пулы' с данными о пулах ликвидности"""

    ws = wb.create_sheet('Пулы', 0)

    # Заголовки столбцов
    headers = [
        'ID Позиции',
        'Пара',
        'Token 0',
        'Token 1',
        'Количество Token 0',
        'Количество Token 1',
        'Цена Token 0 (USD)',
        'Цена Token 1 (USD)',
        'Общая стоимость (USD)',
        'Дата открытия',
        'Статус',
        'Fee Tier (%)',
        'Заработано комиссий (USD)',
        'ROI (%)'
    ]

    # Стили для заголовков
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Применяем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Устанавливаем ширину столбцов
    column_widths = [12, 15, 10, 10, 18, 18, 18, 18, 20, 15, 12, 12, 22, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Добавляем примеры данных
    sample_data = [
        [1, 'ETH/USDC', 'ETH', 'USDC', 2.5, 5000, 2000, 1, '=E2*G2+F2*H2',
         (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'), 'Активна', 0.3, 150, '=(M2/I2)*100'],
        [2, 'WBTC/ETH', 'WBTC', 'ETH', 0.1, 2.0, 40000, 2000, '=E3*G3+F3*H3',
         (datetime.now() - timedelta(days=60)).strftime('%Y-%m-%d'), 'Активна', 0.3, 280, '=(M3/I3)*100'],
        [3, 'USDC/DAI', 'USDC', 'DAI', 10000, 10000, 1, 1, '=E4*G4+F4*H4',
         (datetime.now() - timedelta(days=15)).strftime('%Y-%m-%d'), 'Активна', 0.01, 45, '=(M4/I4)*100'],
        [4, 'ETH/USDT', 'ETH', 'USDT', 1.5, 3000, 2000, 1, '=E5*G5+F5*H5',
         (datetime.now() - timedelta(days=45)).strftime('%Y-%m-%d'), 'Закрыта', 0.3, 95, '=(M5/I5)*100'],
        [5, 'UNI/ETH', 'UNI', 'ETH', 500, 1.0, 8, 2000, '=E6*G6+F6*H6',
         (datetime.now() - timedelta(days=20)).strftime('%Y-%m-%d'), 'Активна', 0.3, 72, '=(M6/I6)*100'],
    ]

    # Добавляем данные
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            if isinstance(value, str) and value.startswith('='):
                cell.value = value
            else:
                cell.value = value

            # Выравнивание
            if col_num in [1, 9, 11, 12, 13, 14]:  # Числовые и процентные столбцы
                cell.alignment = Alignment(horizontal='right')
            elif col_num == 10:  # Дата
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')

            # Форматирование чисел
            if col_num in [7, 8, 9, 13]:  # USD столбцы
                cell.number_format = '$#,##0.00'
            elif col_num in [5, 6]:  # Количество токенов
                cell.number_format = '#,##0.0000'
            elif col_num in [12, 14]:  # Проценты
                cell.number_format = '0.00'

    # Закрепляем первую строку
    ws.freeze_panes = 'A2'

    print("✅ Лист 'Пулы' создан")


def create_profit_analysis_sheet(wb):
    """Создаёт лист 'Анализ доходности' с временными данными"""

    ws = wb.create_sheet('Анализ доходности', 1)

    # Заголовки
    headers = ['Дата', 'Накопленные комиссии (USD)', 'Общая стоимость портфеля (USD)', 'ROI (%)']

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Ширина столбцов
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 12

    # Генерируем временные данные (последние 30 дней)
    start_date = datetime.now() - timedelta(days=30)

    for day in range(31):
        current_date = start_date + timedelta(days=day)
        row = day + 2

        # Дата
        ws.cell(row=row, column=1).value = current_date.strftime('%Y-%m-%d')

        # Накопленные комиссии (растут со временем)
        fees = 50 + (day * 15) + (day % 7) * 10
        ws.cell(row=row, column=2).value = fees
        ws.cell(row=row, column=2).number_format = '$#,##0.00'

        # Общая стоимость портфеля (колеблется)
        portfolio_value = 20000 + (day * 50) - (day % 5) * 200
        ws.cell(row=row, column=3).value = portfolio_value
        ws.cell(row=row, column=3).number_format = '$#,##0.00'

        # ROI
        ws.cell(row=row, column=4).value = f'=(B{row}/C{row})*100'
        ws.cell(row=row, column=4).number_format = '0.00'

    ws.freeze_panes = 'A2'

    print("✅ Лист 'Анализ доходности' создан")


def create_dashboard_sheet(wb):
    """Создаёт главный Dashboard с метриками и графиками"""

    ws = wb.create_sheet('Dashboard', 2)

    # Устанавливаем ширину столбцов для лучшего отображения
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Заголовок Dashboard
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = '📊 Dashboard учёта пулов ликвидности Uniswap'
    title_cell.font = Font(bold=True, size=16, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Стили для метрик
    metric_label_font = Font(bold=True, size=12)
    metric_value_font = Font(bold=True, size=14, color='2E75B6')
    metric_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Ключевые метрики
    current_row = 3

    # 1. Общая стоимость всех позиций
    ws.merge_cells(f'A{current_row}:D{current_row}')
    label_cell = ws[f'A{current_row}']
    label_cell.value = 'Общая стоимость всех позиций'
    label_cell.font = metric_label_font
    label_cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'E{current_row}:H{current_row}')
    value_cell = ws[f'E{current_row}']
    value_cell.value = "=SUM(Пулы!I:I)"
    value_cell.font = metric_value_font
    value_cell.number_format = '$#,##0.00'
    value_cell.alignment = Alignment(horizontal='right', vertical='center')
    value_cell.fill = metric_fill

    for col in range(1, 9):
        ws.cell(row=current_row, column=col).border = border

    ws.row_dimensions[current_row].height = 25
    current_row += 2

    # 2. Всего заработано комиссий
    ws.merge_cells(f'A{current_row}:D{current_row}')
    label_cell = ws[f'A{current_row}']
    label_cell.value = 'Всего заработано комиссий'
    label_cell.font = metric_label_font
    label_cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'E{current_row}:H{current_row}')
    value_cell = ws[f'E{current_row}']
    value_cell.value = "=SUM(Пулы!M:M)"
    value_cell.font = metric_value_font
    value_cell.number_format = '$#,##0.00'
    value_cell.alignment = Alignment(horizontal='right', vertical='center')
    value_cell.fill = metric_fill

    for col in range(1, 9):
        ws.cell(row=current_row, column=col).border = border

    ws.row_dimensions[current_row].height = 25
    current_row += 2

    # 3. Средний ROI
    ws.merge_cells(f'A{current_row}:D{current_row}')
    label_cell = ws[f'A{current_row}']
    label_cell.value = 'Средний ROI'
    label_cell.font = metric_label_font
    label_cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'E{current_row}:H{current_row}')
    value_cell = ws[f'E{current_row}']
    value_cell.value = "=AVERAGE(Пулы!N:N)"
    value_cell.font = metric_value_font
    value_cell.number_format = '0.00"%"'
    value_cell.alignment = Alignment(horizontal='right', vertical='center')
    value_cell.fill = metric_fill

    for col in range(1, 9):
        ws.cell(row=current_row, column=col).border = border

    ws.row_dimensions[current_row].height = 25
    current_row += 2

    # 4. Количество активных позиций
    ws.merge_cells(f'A{current_row}:D{current_row}')
    label_cell = ws[f'A{current_row}']
    label_cell.value = 'Количество активных позиций'
    label_cell.font = metric_label_font
    label_cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'E{current_row}:H{current_row}')
    value_cell = ws[f'E{current_row}']
    value_cell.value = '=COUNTIF(Пулы!K:K,"Активна")'
    value_cell.font = metric_value_font
    value_cell.number_format = '0'
    value_cell.alignment = Alignment(horizontal='right', vertical='center')
    value_cell.fill = metric_fill

    for col in range(1, 9):
        ws.cell(row=current_row, column=col).border = border

    ws.row_dimensions[current_row].height = 25
    current_row += 3

    # График доходности по времени
    chart_row = current_row

    # Заголовок для графика
    ws.merge_cells(f'A{chart_row}:H{chart_row}')
    chart_title = ws[f'A{chart_row}']
    chart_title.value = '📈 График доходности по времени'
    chart_title.font = Font(bold=True, size=12, color='FFFFFF')
    chart_title.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    chart_title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[chart_row].height = 25

    chart_row += 1

    # Создаём линейный график
    line_chart = LineChart()
    line_chart.title = "Накопленные комиссии за период"
    line_chart.style = 12
    line_chart.y_axis.title = 'Комиссии (USD)'
    line_chart.x_axis.title = 'Дата'
    line_chart.width = 20
    line_chart.height = 12

    # Данные для графика из листа "Анализ доходности"
    data = Reference(wb['Анализ доходности'], min_col=2, min_row=1, max_row=32)
    dates = Reference(wb['Анализ доходности'], min_col=1, min_row=2, max_row=32)

    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(dates)

    # Размещаем график
    ws.add_chart(line_chart, f'A{chart_row}')

    chart_row += 20

    # Круговая диаграмма распределения по парам
    ws.merge_cells(f'A{chart_row}:H{chart_row}')
    pie_title = ws[f'A{chart_row}']
    pie_title.value = '🥧 Распределение стоимости по торговым парам'
    pie_title.font = Font(bold=True, size=12, color='FFFFFF')
    pie_title.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    pie_title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[chart_row].height = 25

    chart_row += 1

    # Создаём круговую диаграмму
    pie_chart = PieChart()
    pie_chart.title = "Распределение позиций по парам"
    pie_chart.width = 15
    pie_chart.height = 12

    # Данные для круговой диаграммы
    labels = Reference(wb['Пулы'], min_col=2, min_row=2, max_row=6)
    data = Reference(wb['Пулы'], min_col=9, min_row=1, max_row=6)

    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)

    # Размещаем диаграмму
    ws.add_chart(pie_chart, f'A{chart_row}')

    print("✅ Лист 'Dashboard' создан с метриками и графиками")


if __name__ == '__main__':
    create_uniswap_dashboard()
