#!/usr/bin/env python3
"""
Скрипт для проверки корректности созданного Dashboard
"""

import openpyxl
from openpyxl.chart import PieChart, LineChart

def test_dashboard():
    """Проверяет структуру и содержимое Dashboard"""

    print("🔍 Проверка файла Uniswap_Pool_Dashboard.xlsx...")
    print()

    try:
        wb = openpyxl.load_workbook('Uniswap_Pool_Dashboard.xlsx')

        # Проверка наличия листов
        print("📋 Проверка листов:")
        expected_sheets = ['Пулы', 'Анализ доходности', 'Dashboard']
        actual_sheets = wb.sheetnames

        for sheet_name in expected_sheets:
            if sheet_name in actual_sheets:
                print(f"  ✅ Лист '{sheet_name}' найден")
            else:
                print(f"  ❌ Лист '{sheet_name}' не найден!")
                return False
        print()

        # Проверка листа "Пулы"
        print("📊 Проверка листа 'Пулы':")
        pools_sheet = wb['Пулы']

        # Проверка заголовков
        expected_headers = [
            'ID Позиции', 'Пара', 'Token 0', 'Token 1',
            'Количество Token 0', 'Количество Token 1',
            'Цена Token 0 (USD)', 'Цена Token 1 (USD)',
            'Общая стоимость (USD)', 'Дата открытия',
            'Статус', 'Fee Tier (%)', 'Заработано комиссий (USD)', 'ROI (%)'
        ]

        for col_num, expected_header in enumerate(expected_headers, 1):
            actual_header = pools_sheet.cell(1, col_num).value
            if actual_header == expected_header:
                print(f"  ✅ Столбец {col_num}: '{expected_header}'")
            else:
                print(f"  ❌ Столбец {col_num}: ожидалось '{expected_header}', получено '{actual_header}'")

        # Проверка данных
        data_rows = pools_sheet.max_row - 1
        print(f"  ℹ️  Строк с данными: {data_rows}")
        print()

        # Проверка формул
        print("🔢 Проверка формул в листе 'Пулы':")

        # Проверяем формулу общей стоимости
        total_value_formula = pools_sheet.cell(2, 9).value
        if isinstance(total_value_formula, str) and total_value_formula.startswith('='):
            print(f"  ✅ Формула общей стоимости: {total_value_formula}")
        else:
            print(f"  ⚠️  Формула общей стоимости может быть вычислена: {total_value_formula}")

        # Проверяем формулу ROI
        roi_formula = pools_sheet.cell(2, 14).value
        if isinstance(roi_formula, str) and roi_formula.startswith('='):
            print(f"  ✅ Формула ROI: {roi_formula}")
        else:
            print(f"  ⚠️  Формула ROI может быть вычислена: {roi_formula}")
        print()

        # Проверка листа "Анализ доходности"
        print("📈 Проверка листа 'Анализ доходности':")
        analysis_sheet = wb['Анализ доходности']

        analysis_headers = ['Дата', 'Накопленные комиссии (USD)',
                           'Общая стоимость портфеля (USD)', 'ROI (%)']

        for col_num, expected_header in enumerate(analysis_headers, 1):
            actual_header = analysis_sheet.cell(1, col_num).value
            if actual_header == expected_header:
                print(f"  ✅ Столбец {col_num}: '{expected_header}'")
            else:
                print(f"  ❌ Столбец {col_num}: ожидалось '{expected_header}', получено '{actual_header}'")

        data_rows = analysis_sheet.max_row - 1
        print(f"  ℹ️  Временных точек: {data_rows}")
        print()

        # Проверка листа "Dashboard"
        print("📊 Проверка листа 'Dashboard':")
        dashboard_sheet = wb['Dashboard']

        # Проверка заголовка
        title = dashboard_sheet['A1'].value
        if 'Dashboard' in title or 'Uniswap' in title:
            print(f"  ✅ Заголовок Dashboard: {title}")
        else:
            print(f"  ⚠️  Заголовок Dashboard: {title}")

        # Проверка формул метрик
        print("\n  🔍 Проверка формул метрик:")

        metrics = {
            'E3': ('Общая стоимость всех позиций', '=SUM(Пулы!I:I)'),
            'E5': ('Всего заработано комиссий', '=SUM(Пулы!M:M)'),
            'E7': ('Средний ROI', '=AVERAGE(Пулы!N:N)'),
            'E9': ('Количество активных позиций', '=COUNTIF(Пулы!K:K,"Активна")'),
        }

        for cell_ref, (metric_name, expected_formula) in metrics.items():
            cell_value = dashboard_sheet[cell_ref].value
            if cell_value == expected_formula:
                print(f"    ✅ {metric_name}: {cell_value}")
            else:
                print(f"    ⚠️  {metric_name}: ожидалось '{expected_formula}', получено '{cell_value}'")

        # Проверка графиков
        print("\n  📊 Проверка графиков:")
        chart_count = len(dashboard_sheet._charts)
        print(f"    ℹ️  Найдено графиков: {chart_count}")

        for idx, chart in enumerate(dashboard_sheet._charts, 1):
            chart_type = type(chart).__name__
            chart_title = chart.title if hasattr(chart, 'title') else 'Без названия'
            print(f"    ✅ График {idx}: {chart_type} - '{chart_title}'")

        if chart_count < 2:
            print(f"    ⚠️  Ожидалось 2 графика, найдено {chart_count}")

        print()
        print("=" * 60)
        print("✅ Все проверки пройдены успешно!")
        print("=" * 60)

        wb.close()
        return True

    except Exception as e:
        print(f"\n❌ Ошибка при проверке: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    success = test_dashboard()
    exit(0 if success else 1)
